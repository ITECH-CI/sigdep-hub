#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Génère le dictionnaire de données SIGDEP-3 (schéma core.*) au format Excel."""
import csv, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCHEMA = "schema_core.csv"
OUT = "../SIGDEP3_dictionnaire_donnees.xlsx"

# ---- Descriptions et ordre logique des tables ----
TABLES = [
    ("regions", "Régions sanitaires", "Référentiel géographique — niveau 1 (région). Donnée de référence, seedée, non issue des sites."),
    ("districts", "Districts sanitaires", "Référentiel géographique — niveau 2 (district), rattaché à une région."),
    ("sites", "Sites (établissements de santé)", "Référentiel des établissements. Chaque site a une base OpenMRS locale et un agent de synchronisation. ~3 900 sites."),
    ("identifier_types", "Types d'identifiants patient", "Référentiel des types d'identifiants (UPID national, CODE_ARV, CRUID…). Définit le format attendu de chaque type."),
    ("patients", "Patients", "Un enregistrement patient PAR SITE. Clé d'unicité (site_id, source_uuid). Un même individu suivi sur 2 sites = 2 lignes (voir feuille « Notes importantes »)."),
    ("patient_identifiers", "Identifiants des patients", "Identifiants d'un patient (0..n par patient). L'UPID est l'identifiant national censé être unique inter-sites — mais renseigné dans ~9 % des cas seulement."),
    ("treatment_initiations", "Initiations de traitement ARV", "Mise sous traitement antirétroviral. Une initiation par patient (fiche d'initiation OpenMRS). Contexte clinique et socio-démographique à la mise sous ARV."),
    ("treatment_initiations_pediatric", "Initiations — volet pédiatrique", "Extension pédiatrique d'une initiation (lien 1-1 via initiation_id) : données périnatales, parents/tuteur, prophylaxie du nourrisson."),
    ("visits", "Visites de suivi", "Visites cliniques de suivi ARV. Table PARTITIONNÉE par année de visite. Source des vues Clinique et Pharmacie. La plus volumineuse après lab_results."),
    ("dispensations", "Dispensations ARV (table dédiée)", "Table de dispensation dédiée — NON alimentée actuellement (la dispensation est captée via les visites). Conservée pour évolution."),
    ("lab_results", "Résultats de biologie", "Un résultat d'examen biologique par ligne. La plus volumineuse (~1 M+). Les métadonnées de bilan (unité, prélèvement, n° échantillon) sont exclues depuis l'agent 2.2.4."),
    ("closures", "Clôtures de dossier", "Fin de prise en charge : décès, transfert, arrêt volontaire, sérologie négative. Une clôture par événement de sortie."),
    ("tpt_records", "TPT (thérapie préventive TB)", "Enregistrements de thérapie préventive de la tuberculose (démarrage / suivi / fin)."),
    ("screenings", "Dépistage VIH", "Dépistage VIH (module HIV Screening) : jusqu'à 3 tests, résultat final, population, motif, porte d'entrée."),
    ("ptme_mothers", "PTME — mères", "Prévention de la Transmission Mère-Enfant : suivi des femmes enceintes séropositives."),
    ("ptme_mother_visits", "PTME — visites mères", "Visites de suivi PTME d'une mère."),
    ("ptme_children", "PTME — enfants (nés exposés)", "Suivi des enfants nés de mères séropositives : PCR, sérologies, prophylaxie, résultat de suivi."),
    ("ptme_child_visits", "PTME — visites enfants", "Visites de suivi PTME d'un enfant."),
]
# visits/dispensations absents du CSV information_schema fourni ? -> visits est partitionnée, ses colonnes peuvent ne pas
# apparaître dans le CSV si la requête ciblait le parent. On garde l'entrée descriptive ; les colonnes sortent du CSV si présentes.

# ---- Descriptions par colonne. Clé = (table, colonne). Fallback = COMMON. ----
COMMON = {
    "id": "Identifiant technique interne (clé primaire, auto-incrément). Sans signification métier.",
    "site_id": "Site d'appartenance de l'enregistrement (FK → sites.id).",
    "patient_id": "Patient concerné (FK → patients.id).",
    "source_uuid": "UUID d'origine dans la base OpenMRS du site. Unique PAR SITE (couple site_id, source_uuid).",
    "voided": "Vrai = enregistrement supprimé logiquement (exclu de toutes les statistiques). Défaut : faux.",
    "created_at": "Horodatage d'insertion dans le hub (technique).",
    "updated_at": "Horodatage de dernière mise à jour dans le hub (technique).",
    "extra_data": "Données additionnelles non modélisées, au format JSON (paires concept OpenMRS → valeur). Champ d'extension.",
}
DESC = {
 # regions / districts / sites
 ("regions","code"): "Code de la région (référentiel national).",
 ("regions","name"): "Libellé de la région.",
 ("regions","source_uuid"): "UUID de référence (facultatif).",
 ("districts","code"): "Code du district.",
 ("districts","region_id"): "Région de rattachement (FK → regions.id).",
 ("districts","name"): "Libellé du district.",
 ("sites","code"): "Code établissement (ex. 00064). Correspond à SIGDEP_SITE_CODE côté agent.",
 ("sites","name"): "Nom de l'établissement.",
 ("sites","district_id"): "District de rattachement (FK → districts.id).",
 ("sites","facility_type"): "Type d'établissement (CSU, HG, CHR…).",
 ("sites","latitude"): "Latitude géographique (décimale).",
 ("sites","longitude"): "Longitude géographique (décimale).",
 ("sites","active"): "Site actif dans la plateforme.",
 ("sites","api_key_hash"): "Empreinte (hash) de la clé API du site pour l'ingestion. Technique/sécurité.",
 ("sites","last_sync_at"): "Date-heure de la dernière synchronisation reçue de ce site.",
 ("sites","runs_sigdep"): "Le site fait-il effectivement tourner un agent SIGDEP (vs site référencé sans agent).",
 ("identifier_types","code"): "Code du type d'identifiant (UPID, CODE_ARV, CRUID, UPID…).",
 ("identifier_types","name"): "Libellé du type d'identifiant.",
 ("identifier_types","description"): "Description du type.",
 ("identifier_types","format_regex"): "Expression régulière de validation du format attendu.",
 ("identifier_types","is_active"): "Type d'identifiant actif.",
 # patients
 ("patients","sex"): "Sexe : 'M' (masculin), 'F' (féminin), 'U' (inconnu).",
 ("patients","birth_date"): "Date de naissance.",
 ("patients","birth_date_estimated"): "Vrai si la date de naissance est estimée (jour/mois par défaut).",
 ("patients","birth_place"): "Lieu de naissance (texte libre issu de la source).",
 ("patients","profession"): "Profession (texte).",
 ("patients","education_level"): "Niveau d'éducation.",
 ("patients","marital_status"): "Situation matrimoniale.",
 ("patients","religion"): "Religion.",
 ("patients","voided_at"): "Date-heure de la suppression logique du patient (si voided).",
 # patient_identifiers
 ("patient_identifiers","identifier_type_id"): "Type d'identifiant (FK → identifier_types.id).",
 ("patient_identifiers","identifier_value"): "Valeur de l'identifiant (ex. code ARV, UPID).",
 ("patient_identifiers","is_preferred"): "Identifiant préféré/principal pour ce patient.",
 ("patient_identifiers","valid_from"): "Début de validité (facultatif).",
 ("patient_identifiers","valid_to"): "Fin de validité (facultatif).",
 # treatment_initiations
 ("treatment_initiations","enrollment_date"): "Date d'enrôlement dans la file active. Sert de repli quand arv_init_date est absente.",
 ("treatment_initiations","arv_init_date"): "Date de mise sous ARV. NULL fréquent (~64 %) → les vues « période » retombent sur enrollment_date.",
 ("treatment_initiations","hiv_test_date"): "Date du test VIH.",
 ("treatment_initiations","hiv_type"): "Type de VIH (VIH-1, VIH-2, VIH-1&2).",
 ("treatment_initiations","entry_point"): "Porte d'entrée dans les soins (consultation, PTME, dépistage…).",
 ("treatment_initiations","who_stage_initial"): "Stade clinique OMS à l'initiation (I à IV).",
 ("treatment_initiations","cdc_stage_initial"): "Stade CDC à l'initiation.",
 ("treatment_initiations","arv_regimen_initial"): "Régime ARV initial (ex. TDF 3TC DTG).",
 ("treatment_initiations","weight_initial_kg"): "Poids à l'initiation (kg).",
 ("treatment_initiations","cd4_initial"): "Taux de CD4 à l'initiation (cellules/µL).",
 ("treatment_initiations","cd4_pct_initial"): "Pourcentage de CD4 à l'initiation.",
 ("treatment_initiations","karnofsky_score"): "Indice de Karnofsky (état général, 0–100).",
 ("treatment_initiations","referred"): "Patient référé (Oui/Non).",
 ("treatment_initiations","referred_origin"): "Origine de la référence.",
 ("treatment_initiations","treatment_motive"): "Motif du traitement.",
 ("treatment_initiations","partner_hiv_status"): "Statut VIH du partenaire.",
 ("treatment_initiations","tb_history"): "Antécédent de tuberculose.",
 ("treatment_initiations","arv_history"): "Antécédent de traitement ARV.",
 ("treatment_initiations","transfusion_history"): "Antécédent de transfusion.",
 ("treatment_initiations","ptme_history"): "Antécédent PTME.",
 ("treatment_initiations","ptme_regimen_history"): "Régime PTME antérieur.",
 ("treatment_initiations","ptme_history_date"): "Date de l'antécédent PTME.",
 # pediatric
 ("treatment_initiations_pediatric","initiation_id"): "Initiation parente (FK → treatment_initiations.id, lien 1-1). Clé primaire.",
 ("treatment_initiations_pediatric","birth_weight_kg"): "Poids de naissance (kg).",
 ("treatment_initiations_pediatric","birth_length_cm"): "Taille de naissance (cm).",
 ("treatment_initiations_pediatric","head_circumference_cm"): "Périmètre crânien (cm).",
 ("treatment_initiations_pediatric","apgar_score"): "Score d'Apgar.",
 ("treatment_initiations_pediatric","delivery_mode"): "Mode d'accouchement.",
 ("treatment_initiations_pediatric","delivered_at_facility"): "Accouchement en structure de santé.",
 ("treatment_initiations_pediatric","mother_received_ptme"): "La mère a reçu la PTME.",
 ("treatment_initiations_pediatric","mother_hiv_status"): "Statut VIH de la mère.",
 ("treatment_initiations_pediatric","mother_vital_status"): "Statut vital de la mère.",
 ("treatment_initiations_pediatric","mother_ptme_regimen"): "Régime PTME de la mère.",
 ("treatment_initiations_pediatric","infant_arv_prophylaxis_given"): "Prophylaxie ARV donnée au nourrisson.",
 ("treatment_initiations_pediatric","infant_arv_protocol"): "Protocole ARV du nourrisson.",
 ("treatment_initiations_pediatric","feeding_mode"): "Mode d'alimentation du nourrisson.",
 ("treatment_initiations_pediatric","weaning_date"): "Date de sevrage.",
 ("treatment_initiations_pediatric","vaccinations"): "Vaccinations reçues.",
 ("treatment_initiations_pediatric","schooling_status"): "Statut de scolarisation.",
 ("treatment_initiations_pediatric","admission_date"): "Date d'admission.",
 ("treatment_initiations_pediatric","screening_code"): "Code de dépistage associé.",
 # visits
 ("visits","source_form"): "Formulaire OpenMRS d'origine de la visite (ex. « PEC - Suivi patient »).",
 ("visits","visit_date"): "Date de la visite. Clé de PARTITIONNEMENT de la table. Borne des vues « période ».",
 ("visits","next_visit_date"): "Date du prochain rendez-vous.",
 ("visits","tb_screening_result"): "Résultat du dépistage TB à la visite.",
 ("visits","tb_diagnosed"): "Tuberculose diagnostiquée (booléen).",
 ("visits","tb_treatment_status"): "Statut du traitement TB.",
 ("visits","tb_treatment_start_date"): "Date de début du traitement TB.",
 ("visits","who_stage"): "Stade clinique OMS (1 à 4) à la visite.",
 ("visits","cdc_stage"): "Stade CDC à la visite.",
 ("visits","ctx_prescribed"): "Cotrimoxazole prescrit (booléen).",
 ("visits","ctx_start_date"): "Date de début du cotrimoxazole.",
 ("visits","ivsa_success_confirmation_date"): "IVSA — date de confirmation de succès (suivi des non-stables).",
 ("visits","is_pregnant"): "Patiente enceinte à la visite.",
 ("visits","is_breastfeeding"): "Patiente allaitante à la visite.",
 ("visits","weight_kg"): "Poids (kg).",
 ("visits","height_cm"): "Taille (cm).",
 ("visits","arv_regimen"): "Régime ARV en cours (ex. TDF 3TC DTG). Sert aussi à la vue Pharmacie.",
 ("visits","temperature_c"): "Température (°C).",
 ("visits","pulse"): "Pouls (battements/min).",
 ("visits","respiratory_rate"): "Fréquence respiratoire (cycles/min).",
 ("visits","bp_systolic"): "Tension artérielle systolique (mmHg).",
 ("visits","bp_diastolic"): "Tension artérielle diastolique (mmHg).",
 ("visits","bmi"): "Indice de masse corporelle (calculé).",
 ("visits","mid_upper_arm_circumference"): "Périmètre brachial (cm).",
 ("visits","viral_load"): "Charge virale (copies/mL) relevée à la visite.",
 ("visits","viral_load_date"): "Date de la charge virale.",
 ("visits","cd4_count"): "Taux de CD4 (cellules/µL) relevé à la visite.",
 ("visits","cd4_date"): "Date du CD4.",
 ("visits","arv_treatment_days"): "Nombre de jours d'ARV dispensés (durée de couverture).",
 ("visits","cotrim_treatment_days"): "Nombre de jours de cotrimoxazole dispensés.",
 ("visits","breastfeeding_status"): "Statut d'allaitement.",
 ("visits","tpt_status"): "Statut TPT à la visite.",
 ("visits","tpt_regimen"): "Protocole TPT à la visite.",
 ("visits","ivsa_msd_code"): "IVSA — code MSD (motif/signe).",
 ("visits","ivsa_alert_signs_count"): "IVSA — nombre de signes d'alerte.",
 ("visits","ivsa_neuro_signs_count"): "IVSA — nombre de signes neurologiques.",
 # lab_results
 ("lab_results","encounter_source_uuid"): "UUID de l'encounter (bilan) d'origine — regroupe les résultats d'un même bilan.",
 ("lab_results","test_uuid"): "UUID du concept OpenMRS de l'examen (identifie le type d'examen).",
 ("lab_results","test_name"): "Libellé de l'examen (ex. VIH CHARGE VIRALE, CRÉATININE, CD4).",
 ("lab_results","exam_date"): "Date de l'examen. Indexée pour les vues nationales (idx_lab_exam_date).",
 ("lab_results","value_numeric"): "Résultat numérique (le plus fréquent).",
 ("lab_results","value_text"): "Résultat texte libre (ex. « < LL »).",
 ("lab_results","value_coded"): "Résultat codé : libellé de la réponse (Positif/Négatif/Détectable…). Utilisé pour les examens qualitatifs (Type VIH, HBs, CV qualitative).",
 ("lab_results","unit"): "Unité de mesure du résultat numérique.",
 # closures
 ("closures","closure_type"): "Type de clôture : DEATH (décès), TRANSFER (transfert), et arrêt volontaire / sérologie négative selon les dates renseignées.",
 ("closures","closure_date"): "Date de clôture du dossier.",
 ("closures","transfer_date"): "Date de transfert (si transfert).",
 ("closures","transfer_destination"): "Structure de destination du transfert.",
 ("closures","transfer_reason"): "Motif du transfert.",
 ("closures","death_date"): "Date de décès déclarée.",
 ("closures","actual_death_date"): "Date de décès réelle/constatée.",
 ("closures","death_cause_code"): "Cause de décès (codée).",
 ("closures","death_cause_text"): "Cause de décès (texte libre).",
 ("closures","voluntary_stop_date"): "Date d'arrêt volontaire du traitement.",
 ("closures","hiv_negative_date"): "Date d'infirmation du statut VIH (clôture pour sérologie négative).",
 # tpt
 ("tpt_records","record_type"): "Nature de l'enregistrement TPT (démarrage / suivi / fin).",
 ("tpt_records","record_date"): "Date de l'enregistrement (borne des vues « période »).",
 ("tpt_records","tpt_followup_date"): "Date de suivi TPT.",
 ("tpt_records","tpt_end_date"): "Date de fin du TPT.",
 ("tpt_records","tpt_outcome"): "Issue du TPT (terminé, interrompu…).",
 ("tpt_records","tpt_order_number"): "Numéro d'ordre/prescription TPT.",
 ("tpt_records","adherence"): "Observance au TPT.",
 ("tpt_records","weight_kg"): "Poids (kg) au moment de l'enregistrement.",
 ("tpt_records","next_visit_date"): "Date de la prochaine visite.",
 ("tpt_records","tpt_status"): "Statut du TPT.",
 ("tpt_records","tpt_regimen"): "Protocole TPT (ex. INH, 3HP).",
 # screenings
 ("screenings","screening_code"): "Code du dépistage.",
 ("screenings","screening_date"): "Date du dépistage (borne des vues « période »).",
 ("screenings","result_announcing_date"): "Date d'annonce du résultat.",
 ("screenings","gender"): "Sexe de la personne dépistée.",
 ("screenings","age"): "Âge au dépistage.",
 ("screenings","profession"): "Profession.",
 ("screenings","residence"): "Résidence.",
 ("screenings","marital_status"): "Situation matrimoniale.",
 ("screenings","other_marital_status"): "Précision si « autre » situation matrimoniale.",
 ("screenings","population_type"): "Type de population (générale, clé…).",
 ("screenings","screening_reason"): "Motif du dépistage.",
 ("screenings","other_screening_reason"): "Précision si « autre » motif.",
 ("screenings","test1_reaction"): "Réaction du test 1 (R/NR).",
 ("screenings","test2_reaction"): "Réaction du test 2.",
 ("screenings","test3_reaction"): "Réaction du test 3.",
 ("screenings","test1_invalidated"): "Test 1 invalidé.",
 ("screenings","test2_invalidated"): "Test 2 invalidé.",
 ("screenings","test3_invalidated"): "Test 3 invalidé.",
 ("screenings","final_result"): "Résultat final du dépistage (POS/NEG/IND).",
 ("screenings","retesting"): "Re-test effectué.",
 ("screenings","comment"): "Commentaire libre.",
 ("screenings","screening_site_type"): "Type de site de dépistage (porte d'entrée).",
 ("screenings","screening_post"): "Poste de dépistage.",
 # ptme mothers
 ("ptme_mothers","pregnant_number"): "Numéro de suivi grossesse.",
 ("ptme_mothers","hiv_care_number"): "Numéro de prise en charge VIH.",
 ("ptme_mothers","screening_number"): "Numéro de dépistage.",
 ("ptme_mothers","age"): "Âge de la mère.",
 ("ptme_mothers","marital_status"): "Situation matrimoniale.",
 ("ptme_mothers","spousal_screening"): "Dépistage du conjoint (réalisé/non).",
 ("ptme_mothers","spousal_screening_result"): "Résultat du dépistage conjoint.",
 ("ptme_mothers","start_date"): "Date de début de suivi PTME. Souvent NULL (81 %) → fiches invisibles en vue « période ».",
 ("ptme_mothers","end_date"): "Date de fin de suivi.",
 ("ptme_mothers","arv_status_at_registering"): "Statut ARV à l'inscription.",
 ("ptme_mothers","estimated_delivery_date"): "Date prévue d'accouchement.",
 ("ptme_mothers","pregnancy_outcome"): "Issue de la grossesse.",
 ("ptme_mothers","spousal_screening_date"): "Date du dépistage conjoint.",
 ("ptme_mothers","delivery_type"): "Type d'accouchement.",
 ("ptme_mother_visits","mother_source_uuid"): "UUID source de la mère (rattachement).",
 ("ptme_mother_visits","visit_date"): "Date de la visite.",
 ("ptme_mother_visits","gestational_age"): "Âge gestationnel (semaines).",
 ("ptme_mother_visits","continuing_arv"): "Poursuite des ARV (O/N/NA).",
 ("ptme_mother_visits","continuing_ctx"): "Poursuite du cotrimoxazole (O/N/NA).",
 # ptme children
 ("ptme_children","mother_source_uuid"): "UUID source de la mère (lien enfant → mère).",
 ("ptme_children","child_followup_number"): "Numéro de suivi de l'enfant.",
 ("ptme_children","birth_date"): "Date de naissance de l'enfant.",
 ("ptme_children","gender"): "Sexe de l'enfant.",
 ("ptme_children","arv_prophylaxis_given"): "Prophylaxie ARV donnée (O/N/NA).",
 ("ptme_children","arv_prophylaxis_given_date"): "Date de la prophylaxie ARV.",
 ("ptme_children","followup_end_date"): "Date de fin de suivi.",
 ("ptme_children","pcr1_result"): "Résultat PCR 1 (POS/NEG…).",
 ("ptme_children","pcr2_result"): "Résultat PCR 2.",
 ("ptme_children","pcr3_result"): "Résultat PCR 3.",
 ("ptme_children","hiv_serology1_result"): "Résultat sérologie VIH 1.",
 ("ptme_children","hiv_serology2_result"): "Résultat sérologie VIH 2.",
 ("ptme_children","followup_result"): "Résultat final du suivi de l'enfant.",
 ("ptme_children","followup_result_date"): "Date du résultat final.",
 ("ptme_children","reference_location"): "Lieu de référence.",
 ("ptme_child_visits","child_source_uuid"): "UUID source de l'enfant (rattachement).",
 ("ptme_child_visits","visit_date"): "Date de la visite.",
 ("ptme_child_visits","eating_type"): "Mode d'alimentation.",
 ("ptme_child_visits","modern_contraceptive_method"): "Méthode contraceptive moderne (contexte).",
 ("ptme_child_visits","continuing_ctx"): "Poursuite du cotrimoxazole.",
 ("ptme_child_visits","continuing_inh"): "Poursuite de l'INH (prophylaxie TB).",
}
# Générique pour les colonnes d'âge PTME et dates PCR non listées
def generic_desc(table, col):
    if col.startswith("age_in_day"):   return "Âge de l'enfant en jours à l'événement."
    if col.startswith("age_in_week"):  return "Âge de l'enfant en semaines à l'événement."
    if col.startswith("age_in_month"): return "Âge de l'enfant en mois à l'événement."
    if col.endswith("_sampling_date"): return "Date de prélèvement (" + col.replace("_sampling_date","").upper() + ")."
    if col.endswith("_date") and col.startswith("hiv_serology"): return "Date de la sérologie VIH."
    if col.endswith("_initiation_date"): return "Date d'initiation (" + col.replace("_initiation_date","").upper() + ")."
    for k in ("father_","mother_","guardian_"):
        if col.startswith(k):
            who = {"father_":"père","mother_":"mère","guardian_":"tuteur"}[k]
            rest = col[len(k):]
            m = {"vital_status":"Statut vital du "+who,"education_level":"Niveau d'éducation du "+who,
                 "activity_type":"Type d'activité du "+who,"hiv_status":"Statut VIH du "+who}
            return m.get(rest, "Donnée relative au "+who+".")
    return ""

def describe(table, col):
    if (table,col) in DESC: return DESC[(table,col)]
    if col in COMMON: return COMMON[col]
    g = generic_desc(table, col)
    return g if g else ""

# Concept OpenMRS source par colonne, extrait des constantes _UUID des extracteurs
# sync (commentaire FR = libellé du concept). Clé = (table, colonne) → (uuid, libellé).
# Seulement les mappings DIRECTS et fiables ; colonnes techniques/dérivées → absent.
CONCEPTS = {
 # patients (person_attributes + obs)
 ("patients","marital_status"): ("1054AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","État civil"),
 ("patients","birth_place"): ("164444AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Lieu de naissance"),
 ("patients","education_level"): ("1712AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Niveau d'éducation"),
 ("patients","religion"): ("162894AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Religion"),
 ("patients","profession"): ("162904AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Profession"),
 # treatment_initiations
 ("treatment_initiations","arv_init_date"): ("159599AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date de début ARV"),
 ("treatment_initiations","hiv_test_date"): ("160554AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date diagnostique VIH"),
 ("treatment_initiations","hiv_type"): ("163623AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Résultat Type VIH"),
 ("treatment_initiations","entry_point"): ("164523AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Point d'entrée"),
 ("treatment_initiations","who_stage_initial"): ("164487AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Stade clinique OMS"),
 ("treatment_initiations","cdc_stage_initial"): ("1209AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Stadification CDC"),
 ("treatment_initiations","arv_regimen_initial"): ("162240AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Régime ARV"),
 ("treatment_initiations","weight_initial_kg"): ("5089AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Poids (kg)"),
 ("treatment_initiations","cd4_initial"): ("5497AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Numération CD4"),
 ("treatment_initiations","cd4_pct_initial"): ("730AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","CD4%"),
 ("treatment_initiations","karnofsky_score"): ("5283AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Score de Karnofsky"),
 ("treatment_initiations","referred"): ("1648AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Patient référé"),
 ("treatment_initiations","referred_origin"): ("164562AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Référé origine"),
 ("treatment_initiations","treatment_motive"): ("162225AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Motif mise sous TARV"),
 ("treatment_initiations","partner_hiv_status"): ("1436AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Statut sérologique partenaire"),
 ("treatment_initiations","tb_history"): ("1389AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Antécédent tuberculose"),
 ("treatment_initiations","arv_history"): ("164540AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Antécédent ARV"),
 ("treatment_initiations","transfusion_history"): ("1871AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Antécédent transfusion"),
 ("treatment_initiations","ptme_history"): ("163450AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Antécédent PTME"),
 ("treatment_initiations","ptme_regimen_history"): ("1400AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Régime antécédent PTME"),
 ("treatment_initiations","ptme_history_date"): ("164588AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date PTME"),
 # visits
 ("visits","weight_kg"): ("5089AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Poids (kg)"),
 ("visits","height_cm"): ("5090AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Taille (cm)"),
 ("visits","bmi"): ("1342AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","IMC"),
 ("visits","temperature_c"): ("5088AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Température (°C)"),
 ("visits","pulse"): ("5087AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Pouls"),
 ("visits","respiratory_rate"): ("5242AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Rythme respiratoire"),
 ("visits","bp_systolic"): ("5085AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Tension artérielle systolique"),
 ("visits","bp_diastolic"): ("5086AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Tension artérielle diastolique"),
 ("visits","mid_upper_arm_circumference"): ("163586AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Périmètre brachial"),
 ("visits","who_stage"): ("5356AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Stade clinique OMS (courant)"),
 ("visits","viral_load"): ("856AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","VIH charge virale"),
 ("visits","viral_load_date"): ("165015AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date dernière charge virale"),
 ("visits","cd4_count"): ("159375AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","CD4 (patient reported)"),
 ("visits","cd4_date"): ("160103AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date du dernier CD4"),
 ("visits","arv_regimen"): ("162240AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Régime ARV"),
 ("visits","arv_treatment_days"): ("164590AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Nombre de jours de traitement ARV"),
 ("visits","cotrim_treatment_days"): ("164578AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Nombre de jours de cotrimoxazole"),
 ("visits","next_visit_date"): ("5096AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date de la prochaine visite"),
 ("visits","is_breastfeeding"): ("164764AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Allaitement en cours"),
 ("visits","tpt_status"): ("165049AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Traitement TPT (statut)"),
 ("visits","tpt_regimen"): ("165319AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Protocole TPT"),
 # tpt_records
 ("tpt_records","tpt_followup_date"): ("165234AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date de visite TPT"),
 ("tpt_records","tpt_end_date"): ("165202AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date de fin du TPT"),
 ("tpt_records","tpt_order_number"): ("165244AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Numéro d'ordre TPT"),
 ("tpt_records","tpt_status"): ("165049AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Traitement TPT (statut)"),
 ("tpt_records","tpt_regimen"): ("165319AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Protocole TPT"),
 ("tpt_records","adherence"): ("165200AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Observance traitement préventif"),
 ("tpt_records","weight_kg"): ("5089AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Poids (kg)"),
 # closures (le closure_type est DÉRIVÉ de plusieurs concepts → on liste le principal indicatif)
 ("closures","death_date"): ("1543AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date de décès"),
 ("closures","actual_death_date"): ("165233AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date de décès réelle"),
 ("closures","transfer_date"): ("164595AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Date de transfert"),
 ("closures","transfer_reason"): ("165216AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Motif du transfert"),
 ("closures","transfer_destination"): ("164665AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Destination du transfert"),
 ("closures","death_cause_text"): ("162580AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Cause de décès (texte)"),
 ("closures","death_cause_code"): ("165225AAAAAAAAAAAAAAAAAAAAAAAAAAAAAA","Cause de décès (codée)"),
}
def concept(table, col):
    u = CONCEPTS.get((table,col))
    return f"{u[1]}\n{u[0]}" if u else ""

# Contraintes connues (au-delà de la PK id)
CONSTRAINTS = {
    ("patients","source_uuid"): "Unique (site_id, source_uuid)",
    ("patients","site_id"): "Unique (site_id, source_uuid) • FK sites",
    ("lab_results","source_uuid"): "Unique (site_id, source_uuid)",
    ("closures","source_uuid"): "Unique (site_id, source_uuid)",
    ("treatment_initiations","source_uuid"): "Unique (site_id, source_uuid)",
    ("treatment_initiations_pediatric","initiation_id"): "Clé primaire • FK treatment_initiations",
    ("districts","region_id"): "FK regions",
    ("sites","district_id"): "FK districts",
    ("patient_identifiers","identifier_type_id"): "FK identifier_types",
}
def constraint(table, col):
    if col == "id": return "Clé primaire"
    if (table,col) in CONSTRAINTS: return CONSTRAINTS[(table,col)]
    if col == "patient_id": return "FK patients"
    if col == "site_id": return "FK sites"
    return ""

def clean_type(t):
    # bigint(64,0) -> BIGINT ; character varying(255) -> VARCHAR(255) ; numeric(15,4) -> NUMERIC(15,4)
    t = t.replace("character varying", "VARCHAR").replace("timestamp without time zone","TIMESTAMP")
    t = t.replace("bigint(64,0)","BIGINT").replace("integer(32,0)","INTEGER").replace("smallint(16,0)","SMALLINT")
    t = t.replace("double precision","DOUBLE").replace("boolean","BOOLEAN").replace("date","DATE")
    t = t.replace("uuid","UUID").replace("jsonb","JSONB").replace("text","TEXT")
    return t.upper()

# ---- Lecture du schéma ----
rows = {}
with open(SCHEMA, encoding="utf-8") as f:
    r = csv.DictReader(f, delimiter=";")
    for x in r:
        rows.setdefault(x["table_name"], []).append(x)

# ---- Styles ----
HDR_FILL = PatternFill("solid", fgColor="1F6F63")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F6F63")
SUB_FONT = Font(italic=True, size=10, color="555555")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
PK_FILL = PatternFill("solid", fgColor="FFF3CD")

wb = Workbook()

# ===== Feuille Sommaire =====
ws = wb.active
ws.title = "Sommaire"
ws["A1"] = "SIGDEP-3 — Dictionnaire de données (schéma métier core)"
ws["A1"].font = Font(bold=True, size=16, color="1F6F63")
ws["A2"] = f"Serveur central consolidé des patients vivant avec le VIH — Côte d'Ivoire.  Généré le {datetime.date.today().isoformat()}."
ws["A2"].font = SUB_FONT
ws["A4"] = "Table"; ws["B4"] = "Libellé"; ws["C4"] = "Description"
for c in ("A4","B4","C4"): ws[c].font = HDR_FONT; ws[c].fill = HDR_FILL
r = 5
for tname, label, desc in TABLES:
    ws.cell(r,1,f"core.{tname}").font = Font(bold=True)
    ws.cell(r,2,label)
    ws.cell(r,3,desc).alignment = WRAP
    for col in (1,2,3): ws.cell(r,col).border = BORDER; ws.cell(r,col).alignment = WRAP
    r += 1
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 90
ws.freeze_panes = "A5"

# ===== Feuille Notes importantes =====
wn = wb.create_sheet("Notes importantes")
notes = [
 ("Un patient = une ligne PAR SITE",
  "La clé d'unicité de core.patients est (site_id, source_uuid). Chaque site a sa propre base OpenMRS où le patient "
  "a un UUID différent. Un même individu suivi sur 2 sites apparaît donc en 2 lignes distinctes — le hub ne fusionne pas. "
  "Conséquence : count(*) sur patients compte des ENREGISTREMENTS patient-site, pas des personnes uniques."),
 ("Déduplication inter-sites via l'UPID",
  "Le seul moyen d'identifier un même individu sur plusieurs sites est l'UPID (identifiant national unique), stocké dans "
  "patient_identifiers. Mesure prod : ~9 % des patients ont un UPID renseigné ; parmi eux, très peu de doublons inter-sites "
  "détectés. L'enjeu est l'ADOPTION de l'UPID à la saisie, pas la technique."),
 ("Suppression logique (voided)",
  "Aucune ligne n'est supprimée physiquement : la colonne 'voided' (booléen) marque une suppression logique. "
  "TOUJOURS filtrer 'WHERE voided = FALSE' pour les analyses."),
 ("Dates manquantes fréquentes",
  "Certaines dates métier sont souvent NULL : treatment_initiations.arv_init_date (~64 % — repli sur enrollment_date via "
  "COALESCE), ptme_mothers.start_date (~81 %). Une analyse « par période » sur la date brute exclut ces lignes."),
 ("extra_data (JSON)",
  "La colonne extra_data (JSONB) porte des données non modélisées : paires 'UUID de concept OpenMRS → valeur'. "
  "À exploiter au cas par cas ; les clés sont des UUID, pas des libellés."),
 ("lab_results — métadonnées exclues",
  "Depuis l'agent 2.2.4, les obs non-résultat (unité de mesure, type de prélèvement, n° d'échantillon, grossesse/allaitement) "
  "ne sont plus dans lab_results. Ces métadonnées de contexte pourront être re-modélisées ultérieurement."),
 ("Partitionnement de visits",
  "core.visits est partitionnée par année (visit_date), avec une partition DEFAULT et des partitions jusqu'en 2040. "
  "Améliore les performances (partition pruning) et évite tout rejet de date hors plage."),
 ("Sites, régions, districts = référentiels",
  "regions / districts / sites / identifier_types sont des données de RÉFÉRENCE (seedées), pas issues de la synchronisation "
  "des patients. Elles servent à filtrer/agréger géographiquement."),
]
wn["A1"] = "Notes importantes pour l'analyse des données"
wn["A1"].font = TITLE_FONT
r = 3
for title, body in notes:
    wn.cell(r,1,title).font = Font(bold=True, size=11, color="1F6F63")
    wn.cell(r,1).alignment = WRAP
    wn.cell(r+1,1,body).alignment = WRAP
    r += 3
wn.column_dimensions["A"].width = 120

# ===== Une feuille par table =====
HEADERS = ["#","Colonne","Type SQL","Obligatoire","Clé / Contrainte","Description métier","Concept OpenMRS (libellé + UUID)"]
WIDTHS  = [5, 30, 18, 11, 28, 60, 34]
label_by = {t: lbl for t,lbl,_ in TABLES}
desc_by  = {t: d   for t,_,d   in TABLES}

for tname, label, tdesc in TABLES:
    cols = rows.get(tname)
    if not cols:  # ex. visits/dispensations absents du CSV
        continue
    title = tname[:28]  # nom d'onglet <= 31 chars
    wt = wb.create_sheet(title)
    wt.cell(1,1,f"core.{tname} — {label}").font = TITLE_FONT
    wt.cell(2,1,tdesc).font = SUB_FONT
    wt.cell(2,1).alignment = WRAP
    wt.merge_cells(start_row=2,start_column=1,end_row=2,end_column=7)
    hr = 4
    for i,h in enumerate(HEADERS, start=1):
        c = wt.cell(hr,i,h); c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER; c.alignment = Alignment(vertical="center")
    rr = hr+1
    for x in sorted(cols, key=lambda z:int(z["ordinal_position"])):
        col = x["column_name"]
        vals = [x["ordinal_position"], col, clean_type(x["type_sql"]),
                "Oui" if x["is_nullable"]=="NO" else "", constraint(tname,col), describe(tname,col),
                concept(tname,col)]
        for i,v in enumerate(vals, start=1):
            c = wt.cell(rr,i,v); c.border = BORDER; c.alignment = WRAP
            if constraint(tname,col)=="Clé primaire" or "Unique" in constraint(tname,col):
                c.fill = PK_FILL
        rr += 1
    for i,w in enumerate(WIDTHS, start=1):
        wt.column_dimensions[get_column_letter(i)].width = w
    wt.freeze_panes = "A5"

wb.save(OUT)
print("OK:", OUT)
print("Feuilles:", wb.sheetnames)
